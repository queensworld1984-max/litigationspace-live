"""
Document Processor Service — handles multi-document upload, text extraction,
OCR for images/scanned docs, and voice transcription via OpenAI Whisper.

Supports: All file types. Text extraction for PDF, DOCX, TXT, RTF, CSV, Excel (XLSX/XLS), images (OCR), audio.
Max: 20 documents, 100MB total
"""
import os
import re
import io
import logging
import base64
import tempfile
from typing import Optional

import httpx

logger = logging.getLogger("document_processor")

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

# Supported file types
TEXT_EXTENSIONS = {"txt", "rtf", "csv", "tsv", "md", "json", "xml", "html", "htm"}
PDF_EXTENSIONS = {"pdf"}
DOCX_EXTENSIONS = {"docx", "doc"}
SPREADSHEET_EXTENSIONS = {"xlsx", "xls", "ods"}
IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "tiff", "tif", "bmp", "webp"}
AUDIO_EXTENSIONS = {"webm", "mp3", "wav", "m4a", "ogg", "mp4"}

ALL_SUPPORTED = TEXT_EXTENSIONS | PDF_EXTENSIONS | DOCX_EXTENSIONS | SPREADSHEET_EXTENSIONS | IMAGE_EXTENSIONS | AUDIO_EXTENSIONS

MAX_FILES = 20
MAX_TOTAL_SIZE = 100 * 1024 * 1024  # 100MB


def get_file_extension(filename: str) -> str:
    """Get lowercase file extension."""
    return filename.lower().rsplit(".", 1)[-1] if "." in filename else ""


def is_supported_file(filename: str) -> bool:
    """Accept all file types — text extraction may be limited for uncommon formats."""
    return True


def is_image_file(filename: str) -> bool:
    """Check if file is an image."""
    return get_file_extension(filename) in IMAGE_EXTENSIONS


def is_audio_file(filename: str) -> bool:
    """Check if file is audio."""
    return get_file_extension(filename) in AUDIO_EXTENSIONS


def extract_text_from_bytes(content: bytes, filename: str) -> str:
    """Extract text from file content based on file type."""
    ext = get_file_extension(filename)

    if ext in TEXT_EXTENSIONS:
        return _extract_text_file(content)
    elif ext in PDF_EXTENSIONS:
        return _extract_pdf_text(content)
    elif ext in DOCX_EXTENSIONS:
        return _extract_docx_text(content, filename)
    elif ext in SPREADSHEET_EXTENSIONS:
        return _extract_spreadsheet_text(content, filename)
    elif ext in IMAGE_EXTENSIONS:
        return _extract_image_text(content, filename)
    else:
        # For any other file type, try reading as text; otherwise store as attachment
        try:
            text = content.decode("utf-8")
            if text.strip():
                return text
        except (UnicodeDecodeError, ValueError):
            pass
        try:
            text = content.decode("latin-1")
            if text.strip() and len([c for c in text[:200] if c.isprintable() or c in '\n\r\t']) > len(text[:200]) * 0.7:
                return text
        except Exception:
            pass
        return f"[File uploaded: {filename}. This file type (.{ext}) does not support automatic text extraction.]"


def _extract_text_file(content: bytes) -> str:
    """Extract text from plain text files."""
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def _extract_pdf_text(content: bytes) -> str:
    """Extract text from PDF using PyMuPDF (fitz) for better quality."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=content, filetype="pdf")
        text_parts = []
        for page in doc:
            text = page.get_text()
            if text.strip():
                text_parts.append(text.strip())
        doc.close()
        if text_parts:
            return "\n\n".join(text_parts)
    except Exception as e:
        logger.warning(f"PyMuPDF extraction failed: {e}")

    # Fallback: try PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(content))
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text and text.strip():
                text_parts.append(text.strip())
        if text_parts:
            return "\n\n".join(text_parts)
    except Exception as e:
        logger.warning(f"PyPDF2 extraction failed: {e}")

    # If no text extracted, it might be a scanned PDF — try OCR via image extraction
    try:
        return _ocr_pdf_pages(content)
    except Exception as e:
        logger.warning(f"PDF OCR fallback failed: {e}")

    return "[Could not extract text from PDF. The document may be scanned or image-only.]"


def _ocr_pdf_pages(content: bytes) -> str:
    """Extract images from PDF pages and OCR them using OpenAI Vision."""
    import fitz
    doc = fitz.open(stream=content, filetype="pdf")
    all_text = []

    for page_num, page in enumerate(doc):
        # Render page to image
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes("png")

        # Use OpenAI Vision for OCR
        text = _ocr_image_with_openai(img_bytes, f"page_{page_num + 1}.png")
        if text and text.strip():
            all_text.append(f"--- Page {page_num + 1} ---\n{text.strip()}")

        # Limit to first 20 pages for performance
        if page_num >= 19:
            all_text.append(f"[... {len(doc) - 20} additional pages not processed]")
            break

    doc.close()
    return "\n\n".join(all_text) if all_text else "[No text could be extracted from scanned PDF.]"


def _extract_docx_text(content: bytes, filename: str = "document.docx") -> str:
    """Extract text from DOCX/DOC using python-docx, LibreOffice, or ZIP/XML fallback."""
    ext = get_file_extension(filename)

    # For .doc (old binary format), use LibreOffice to convert
    if ext == "doc":
        return _extract_doc_with_libreoffice(content, filename)

    try:
        from docx import Document
        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        if paragraphs:
            return "\n\n".join(paragraphs)
    except Exception:
        pass

    # Fallback: ZIP-based extraction
    try:
        import zipfile
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            if "word/document.xml" in zf.namelist():
                xml_content = zf.read("word/document.xml").decode("utf-8")
                texts = re.findall(r"<w:t[^>]*>([^<]+)</w:t>", xml_content)
                return " ".join(texts)
    except Exception:
        pass

    # Last resort: try LibreOffice
    return _extract_doc_with_libreoffice(content, filename)


def _extract_doc_with_libreoffice(content: bytes, filename: str) -> str:
    """Convert .doc/.docx to text using LibreOffice."""
    import subprocess
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, filename)
            with open(input_path, "wb") as f:
                f.write(content)
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "txt:Text", "--outdir", tmpdir, input_path],
                capture_output=True, timeout=30
            )
            txt_path = os.path.join(tmpdir, os.path.splitext(filename)[0] + ".txt")
            if os.path.exists(txt_path):
                with open(txt_path, "r", encoding="utf-8", errors="replace") as f:
                    text = f.read()
                if text.strip():
                    return text.strip()
    except Exception as e:
        logger.warning(f"LibreOffice conversion failed for {filename}: {e}")
    return "[Could not extract text from document.]"


def _extract_spreadsheet_text(content: bytes, filename: str) -> str:
    """Extract text from Excel/spreadsheet files (xlsx, xls, ods)."""
    ext = get_file_extension(filename)

    # Try openpyxl for xlsx
    if ext == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            all_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append(" | ".join(cells))
                if rows:
                    all_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
            wb.close()
            if all_text:
                return "\n\n".join(all_text)
        except Exception as e:
            logger.warning(f"openpyxl extraction failed: {e}")

    # Try xlrd for xls (legacy Excel)
    if ext == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            all_text = []
            for sheet in wb.sheets():
                rows = []
                for row_idx in range(sheet.nrows):
                    cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
                    if any(cells):
                        rows.append(" | ".join(cells))
                if rows:
                    all_text.append(f"--- Sheet: {sheet.name} ---\n" + "\n".join(rows))
            if all_text:
                return "\n\n".join(all_text)
        except Exception as e:
            logger.warning(f"xlrd extraction failed: {e}")

    # Fallback: try ZIP-based extraction for xlsx/ods (they are ZIP archives)
    try:
        import zipfile
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                if name.endswith(".xml") or name.endswith(".csv"):
                    xml_content = zf.read(name).decode("utf-8", errors="replace")
                    texts = re.findall(r">([^<]+)<", xml_content)
                    clean = [t.strip() for t in texts if t.strip() and len(t.strip()) > 0]
                    if clean:
                        return "\n".join(clean[:500])
    except Exception:
        pass

    return f"[Spreadsheet file uploaded: {filename}. Text extraction not available for this format.]"


def _extract_image_text(content: bytes, filename: str) -> str:
    """Extract text from images using OpenAI Vision API, with tesseract fallback."""
    text = _ocr_image_with_openai(content, filename)
    if text and not text.startswith("[OCR failed") and not text.startswith("[Cannot process"):
        return text
    # Fallback to local tesseract OCR
    return _ocr_image_with_tesseract(content, filename)


def _ocr_image_with_openai(image_bytes: bytes, filename: str) -> str:
    """Use OpenAI Vision API to extract text from an image."""
    if not OPENAI_API_KEY:
        return "[Cannot process image: OpenAI API key not configured for OCR.]"

    b64 = base64.b64encode(image_bytes).decode("utf-8")

    # Determine MIME type
    ext = get_file_extension(filename)
    mime_map = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png", "tiff": "image/tiff",
        "tif": "image/tiff", "bmp": "image/bmp",
        "webp": "image/webp",
    }
    mime = mime_map.get(ext, "image/png")

    try:
        with httpx.Client(timeout=60.0) as client:
            resp = client.post(
                "https://api.openai.com/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {OPENAI_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "gpt-5.4",
                    "messages": [
                        {
                            "role": "system",
                            "content": "You are a precise document OCR system. Extract ALL text from the image exactly as it appears. Preserve formatting, paragraphs, dates, numbers, and names. If the image contains a legal document, capture every word. If it contains a handwritten document, do your best to read it. Output only the extracted text, nothing else.",
                        },
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": "Extract all text from this document image:"},
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:{mime};base64,{b64}",
                                        "detail": "high",
                                    },
                                },
                            ],
                        },
                    ],
                    "max_completion_tokens": 8000,
                    "temperature": 0.1,
                },
            )
            if resp.status_code == 200:
                data = resp.json()
                return data["choices"][0]["message"]["content"]
            else:
                logger.error(f"OpenAI Vision error {resp.status_code}: {resp.text[:300]}")
                return f"[OCR failed: API error {resp.status_code}]"
    except Exception as e:
        logger.error(f"OCR request failed: {e}")
        return f"[OCR failed: {str(e)}]"


def _ocr_image_with_tesseract(image_bytes: bytes, filename: str) -> str:
    """Fallback OCR using local tesseract."""
    try:
        from PIL import Image
        import pytesseract
        img = Image.open(io.BytesIO(image_bytes))
        text = pytesseract.image_to_string(img)
        if text and text.strip():
            return text.strip()
    except Exception as e:
        logger.warning(f"Tesseract OCR failed for {filename}: {e}")
    return f"[Could not extract text from image: {filename}]"


async def transcribe_audio(audio_bytes: bytes, filename: str) -> str:
    """Transcribe audio using OpenAI Whisper API."""
    if not OPENAI_API_KEY:
        return "[Cannot transcribe: OpenAI API key not configured.]"

    ext = get_file_extension(filename)
    mime_map = {
        "webm": "audio/webm", "mp3": "audio/mpeg",
        "wav": "audio/wav", "m4a": "audio/mp4",
        "ogg": "audio/ogg", "mp4": "audio/mp4",
    }
    content_type = mime_map.get(ext, "audio/webm")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            files = {
                "file": (filename, audio_bytes, content_type),
                "model": (None, "whisper-1"),
                "language": (None, "en"),
                "response_format": (None, "text"),
            }
            resp = await client.post(
                "https://api.openai.com/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
                files=files,
            )
            if resp.status_code == 200:
                return resp.text.strip()
            else:
                logger.error(f"Whisper error {resp.status_code}: {resp.text[:300]}")
                return f"[Transcription failed: API error {resp.status_code}]"
    except Exception as e:
        logger.error(f"Whisper request failed: {e}")
        return f"[Transcription failed: {str(e)}]"
