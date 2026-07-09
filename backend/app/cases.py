"""Cases router - CRUD operations with multi-tenant isolation."""
import json
import io
import logging
import os
import shutil
import traceback
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional
from pydantic import BaseModel
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import FileResponse

from app.database import get_db
from app.models.schemas import CaseCreate, CaseUpdate, TaskCreate, TaskUpdate, DocumentCreate
from app.utils.auth import get_current_user, generate_id
from app.utils.model_router import get_model_for_task

logger = logging.getLogger(__name__)

# File upload configuration
UPLOAD_BASE_DIR = os.environ.get("UPLOAD_DIR", "/var/www/litigationspace/data/uploads")
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
ALLOWED_EXTENSIONS = {".pdf", ".docx", ".doc", ".txt", ".rtf", ".odt", ".png", ".jpg", ".jpeg", ".gif", ".webp", ".xlsx", ".csv", ".xls", ".pptx", ".ppt", ".tiff", ".tif", ".bmp", ".svg", ".heic", ".heif", ".msg", ".eml", ".pages", ".numbers", ".key"}


def _add_file_to_pdf(merged_pdf, fp: Path, doc_label: str = ""):
    """Convert any supported file type to PDF pages and append to merged_pdf.
    Supports: PDF, images, DOCX/DOC, TXT, RTF, CSV, XLSX."""
    import fitz
    ext = fp.suffix.lower()
    try:
        if ext == ".pdf":
            src = fitz.open(str(fp))
            merged_pdf.insert_pdf(src)
            src.close()
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            img_doc = fitz.open()
            img_page = img_doc.new_page(width=612, height=792)
            img_rect = fitz.Rect(36, 36, 576, 756)
            img_page.insert_image(img_rect, filename=str(fp))
            merged_pdf.insert_pdf(img_doc)
            img_doc.close()
        elif ext in (".docx", ".doc"):
            doc_bytes = fp.read_bytes()
            text = _extract_text_from_docx(doc_bytes)
            if text:
                _add_text_pages(merged_pdf, text, doc_label)
        elif ext in (".txt", ".rtf", ".odt"):
            try:
                text = fp.read_text(encoding="utf-8", errors="replace")
            except Exception:
                text = fp.read_bytes().decode("utf-8", errors="replace")
            if text.strip():
                _add_text_pages(merged_pdf, text.strip(), doc_label)
        elif ext == ".csv":
            text = fp.read_text(encoding="utf-8", errors="replace")
            if text.strip():
                _add_text_pages(merged_pdf, text.strip(), doc_label)
        elif ext == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(fp), data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    lines.append(f"--- Sheet: {sheet} ---")
                    for row in ws.iter_rows(values_only=True):
                        lines.append("\t".join(str(c) if c is not None else "" for c in row))
                text = "\n".join(lines)
                if text.strip():
                    _add_text_pages(merged_pdf, text.strip(), doc_label)
            except ImportError:
                text = f"[XLSX file: {fp.name} — openpyxl not available]"
                _add_text_pages(merged_pdf, text, doc_label)
        else:
            # Unsupported — add a placeholder page
            _add_text_pages(merged_pdf, f"[File: {fp.name} — format not directly convertible]", doc_label)
    except Exception as e:
        logger.warning(f"_add_file_to_pdf failed for {fp}: {e}")


def _add_text_pages(merged_pdf, text: str, title: str = ""):
    """Add text content as PDF pages, handling multi-page overflow."""
    import fitz
    # Split text into chunks that fit on a page
    max_chars_per_page = 3000
    chunks = []
    while text:
        chunks.append(text[:max_chars_per_page])
        text = text[max_chars_per_page:]
    for i, chunk in enumerate(chunks):
        page = merged_pdf.new_page(width=612, height=792)
        # Add title on first page
        if title and i == 0:
            title_rect = fitz.Rect(72, 40, 540, 65)
            page.insert_textbox(title_rect, title, fontsize=12, fontname="helv-b" if hasattr(fitz, 'helv-b') else "helv")
            text_rect = fitz.Rect(72, 72, 540, 750)
        else:
            text_rect = fitz.Rect(72, 50, 540, 750)
        page.insert_textbox(text_rect, chunk, fontsize=10, fontname="helv")


def _apply_bates_numbering(merged_pdf, prefix: str, start_num: int = 1):
    """Stamp Bates numbers on every page of a merged PDF."""
    import fitz
    for page_idx in range(merged_pdf.page_count):
        page = merged_pdf[page_idx]
        bates_num = start_num + page_idx
        bates_text = f"{prefix}-{bates_num:06d}"
        # Position: bottom-right corner
        rect = fitz.Rect(440, 760, 590, 780)
        # White background for readability
        page.draw_rect(rect, color=(0.6, 0.6, 0.6), fill=(1, 1, 1), width=0.5)
        page.insert_textbox(
            fitz.Rect(442, 762, 588, 778),
            bates_text,
            fontsize=8,
            fontname="cour",
            color=(0.2, 0.2, 0.2),
            align=2  # right align
        )


# --------------- Exhibit helpers ---------------

def _exhibit_letter(n: int) -> str:
    """Convert 0-based index to exhibit letter: 0->A, 25->Z, 26->AA, 27->AB, etc."""
    if n < 26:
        return chr(65 + n)
    result = ""
    while n >= 0:
        result = chr(65 + (n % 26)) + result
        n = n // 26 - 1
    return result


def _next_exhibit_label(db, case_id: str, numbering: str = "letters") -> tuple:
    """Return (next_label, next_order) for a case's documents.
    numbering: 'letters' -> A, B, C...  'numbers' -> 1, 2, 3..."""
    row = db.execute(
        "SELECT COUNT(*) as cnt FROM documents WHERE case_id = ? AND exhibit_label IS NOT NULL",
        (case_id,)
    ).fetchone()
    count = row["cnt"] if row else 0
    if numbering == "numbers":
        return str(count + 1), count
    return _exhibit_letter(count), count


def _extract_text_from_pdf(file_bytes: bytes, max_chars: int = 3000) -> str:
    """Extract text from PDF bytes using PyMuPDF."""
    try:
        import fitz
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text()
            if len(text) >= max_chars:
                break
        doc.close()
        return text[:max_chars]
    except Exception as e:
        logger.warning(f"PDF text extraction failed: {e}")
        return ""


def _extract_text_from_docx(file_bytes: bytes, max_chars: int = 3000) -> str:
    """Extract text from DOCX bytes."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = tree.findall(".//w:p", ns)
        text = "\n".join(
            "".join(node.text or "" for node in p.findall(".//w:t", ns))
            for p in paragraphs
        )
        return text[:max_chars]
    except Exception as e:
        logger.warning(f"DOCX text extraction failed: {e}")
        return ""


def _ai_describe_image(file_bytes: bytes, original_filename: str, mime_type: str = "image/png") -> str:
    """Use GPT-4o vision to describe an image for exhibit naming."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return ""
    try:
        import base64
        from openai import OpenAI
        b64 = base64.b64encode(file_bytes).decode("utf-8")
        # Map common extensions to mime types
        ext = Path(original_filename).suffix.lower()
        mime_map = {
            ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
            ".tiff": "image/tiff", ".tif": "image/tiff",
        }
        img_mime = mime_map.get(ext, mime_type)
        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model=get_model_for_task("image_description"),
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a legal document analyst. Describe this image/screenshot in detail. "
                        "Focus on: any text visible, document type, parties mentioned, dates, "
                        "signatures, headers, or notable content. This will be used to generate "
                        "an exhibit name for a legal case. Be specific and factual."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": f"Describe this image (filename: {original_filename}). What does it show?"},
                        {"type": "image_url", "image_url": {"url": f"data:{img_mime};base64,{b64}", "detail": "high"}},
                    ],
                },
            ],
            max_tokens=500,
            temperature=0.3,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logger.warning(f"AI image description failed: {e}")
        return ""


def _ai_generate_case_tasks(jurisdiction: str, forum: str, matter_type: str, case_title: str) -> list:
    """Use OpenAI to generate practical, procedural tasks for a new case.
    Returns a list of task title strings, or empty list if AI is unavailable."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return []
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        prompt = (
            f"You are an expert legal case manager. Generate a practical, ordered task list for the following matter:\n\n"
            f"Jurisdiction: {jurisdiction or 'Unknown'}\n"
            f"Forum / Agency / Court: {forum or 'Unknown'}\n"
            f"Matter Type: {matter_type or 'Unknown'}\n"
            f"Case Title: {case_title}\n\n"
            f"Rules:\n"
            f"- Return ONLY a JSON array of strings, no commentary, no markdown, no extra text.\n"
            f"- Each string is a single task title (max 120 chars).\n"
            f"- Generate 15-25 tasks.\n"
            f"- Tasks must be specific to this jurisdiction, forum, and matter type.\n"
            f"- Order tasks chronologically by when they typically occur in the proceeding.\n"
            f"- Use the correct procedural terminology for this forum (e.g. USCIS vs EOIR vs BIA vs High Court).\n"
            f"- Do NOT include generic tasks like 'Open file' or 'Create folder'.\n"
            f"- Include deadlines, filings, client prep, government response tracking, hearings, and post-decision steps.\n\n"
            f"Return only the JSON array."
        )
        response = client.chat.completions.create(
            model=get_model_for_task("case_task_generation"),
            messages=[{"role": "user", "content": prompt}],
            max_completion_tokens=1200,
            temperature=0.3,
        )
        raw = response.choices[0].message.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        tasks = json.loads(raw.strip())
        if isinstance(tasks, list):
            return [str(t) for t in tasks if t]
        return []
    except Exception as e:
        logger.warning(f"[AI TASKS] Generation failed: {e}")
        return []


def _ai_generate_exhibit_name(text: str, original_filename: str) -> str:
    """Call OpenAI to generate a descriptive exhibit name from document text."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key or not text.strip():
        # Fallback: use cleaned filename
        name = Path(original_filename).stem.replace("_", " ").replace("-", " ")
        return name

    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model=get_model_for_task("exhibit_naming"),
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a legal document analyst. Given the text of a legal document, "
                        "generate a short, descriptive exhibit name. Format: a concise description "
                        "of what the document is, including key parties, dates, and type. "
                        "Examples: 'ERC Consulting Services Agreement dated May 17, 2023 (signed by Derek Cowan)', "
                        "'Invoice #1234 from ABC Corp dated June 3, 2023', "
                        "'Email correspondence between Smith and Jones re: payment terms (Jan 5-12, 2024)'. "
                        "If no date is found in the document, simply omit the date — do NOT write "
                        "'date not found', 'undated', or any similar placeholder. "
                        "Return ONLY the exhibit name, nothing else. Keep it under 120 characters."
                    ),
                },
                {
                    "role": "user",
                    "content": f"Original filename: {original_filename}\n\nDocument text (first ~3000 chars):\n{text}",
                },
            ],
            max_tokens=150,
            temperature=0.3,
        )
        name = response.choices[0].message.content.strip().strip('"').strip("'")
        return name if name else Path(original_filename).stem.replace("_", " ")
    except Exception as e:
        logger.warning(f"AI exhibit naming failed: {e}")
        return Path(original_filename).stem.replace("_", " ").replace("-", " ")


def _ai_generate_document_name(text: str, original_filename: str) -> str:
    """Call OpenAI to generate a clean descriptive name for a non-exhibit document."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key or not text.strip():
        name = Path(original_filename).stem.replace("_", " ").replace("-", " ")
        return name

    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model=get_model_for_task("document_naming"),
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a legal document analyst. Given the text of a document, "
                        "generate a short, descriptive filename. Describe what the document is — "
                        "include key parties and dates if present. "
                        "If no date is found, simply omit the date — do NOT write "
                        "'date not found', 'undated', or any similar placeholder. "
                        "Examples: 'Motion to Dismiss — Smith v. Jones', "
                        "'Retainer Agreement with Johnson LLC dated March 2024', "
                        "'Court Order Granting Summary Judgment'. "
                        "Return ONLY the document name, nothing else. Keep it under 120 characters."
                    ),
                },
                {
                    "role": "user",
                    "content": f"Original filename: {original_filename}\n\nDocument text (first ~3000 chars):\n{text}",
                },
            ],
            max_tokens=150,
            temperature=0.3,
        )
        name = response.choices[0].message.content.strip().strip('"').strip("'")
        return name if name else Path(original_filename).stem.replace("_", " ")
    except Exception as e:
        logger.warning(f"AI document naming failed: {e}")
        return Path(original_filename).stem.replace("_", " ").replace("-", " ")


def _stamp_exhibit_on_pdf(file_bytes: bytes, exhibit_label: str, exhibit_name: str) -> bytes:
    """Insert a full exhibit cover page BEFORE the original document.
    Cover page layout (matching user's template):
      - Vertically centered content on a full page
      - 'EXHIBIT A' in very large bold red text (48pt)
      - '(Document Name)' in bold black text (28pt), in parentheses
    Original document pages follow unchanged."""
    try:
        import fitz
        doc = fitz.open(stream=file_bytes, filetype="pdf")

        if len(doc) == 0:
            doc.close()
            return file_bytes

        # Use same dimensions as first page of original document
        first_page = doc[0]
        pw = first_page.rect.width
        ph = first_page.rect.height

        # Create new document: cover page + all original pages
        new_doc = fitz.open()

        # --- Cover Page ---
        cover = new_doc.new_page(width=pw, height=ph)

        # "EXHIBIT A" — large red, centered horizontally and ~35% from top
        exhibit_text = f"EXHIBIT {exhibit_label}"
        font = fitz.Font("helv")
        text_width = font.text_length(exhibit_text, fontsize=48)
        x_exhibit = (pw - text_width) / 2
        y_exhibit = ph * 0.38
        cover.insert_text(
            fitz.Point(x_exhibit, y_exhibit),
            exhibit_text,
            fontsize=48,
            fontname="helv",
            color=(1, 0, 0),  # RED
        )

        # "(Document Name)" — bold black, centered below exhibit label
        # Wrap long names into multiple lines
        if exhibit_name:
            name_fontsize = 28
            max_width = pw - 80  # 40px margin each side
            name_text = f"({exhibit_name})"

            # Simple word-wrap
            words = name_text.split()
            lines = []
            current_line = ""
            for word in words:
                test = f"{current_line} {word}".strip() if current_line else word
                if font.text_length(test, fontsize=name_fontsize) <= max_width:
                    current_line = test
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word
            if current_line:
                lines.append(current_line)

            y_name = y_exhibit + 60  # gap below exhibit text
            for line in lines:
                line_width = font.text_length(line, fontsize=name_fontsize)
                x_name = (pw - line_width) / 2
                cover.insert_text(
                    fitz.Point(x_name, y_name),
                    line,
                    fontsize=name_fontsize,
                    fontname="helv",
                    color=(0, 0, 0),  # black
                )
                y_name += name_fontsize + 8  # line spacing

        # --- Copy all original pages unchanged ---
        for i in range(len(doc)):
            orig_page = doc[i]
            opw = orig_page.rect.width
            oph = orig_page.rect.height
            np = new_doc.new_page(width=opw, height=oph)
            np.show_pdf_page(orig_page.rect, doc, i)

        output = new_doc.tobytes()
        new_doc.close()
        doc.close()
        return output
    except Exception as e:
        logger.warning(f"PDF stamping failed: {e}")
        return file_bytes  # Return original if stamping fails


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".tif", ".heic", ".heif"}


def _process_exhibit_async(doc_id: str, case_id: str, tenant_id: str,
                           file_bytes: bytes, original_filename: str, ext: str,
                           file_path_on_disk: Path, relative_path: str,
                           category: str = "evidence"):
    """Background task: extract text, AI name, stamp PDF, update DB.
    Only documents with category 'evidence' get exhibit labels and cover pages.
    All other categories just get an AI-generated descriptive filename."""
    try:
        # Look up case exhibit_numbering preference
        numbering = "letters"
        try:
            with get_db() as db:
                case_row = db.execute("SELECT exhibit_numbering FROM cases WHERE id = ?", (case_id,)).fetchone()
                if case_row and case_row["exhibit_numbering"]:
                    numbering = case_row["exhibit_numbering"]
        except Exception:
            pass  # column may not exist yet; default to letters

        # 1. Extract text (or describe image with vision AI)
        if ext == ".pdf":
            text = _extract_text_from_pdf(file_bytes)
        elif ext in (".docx", ".doc"):
            text = _extract_text_from_docx(file_bytes)
        elif ext == ".txt":
            text = file_bytes.decode("utf-8", errors="ignore")[:3000]
        elif ext in IMAGE_EXTENSIONS:
            # Use GPT-4o vision to read the image content
            text = _ai_describe_image(file_bytes, original_filename)
        else:
            text = ""

        is_evidence = (category == "evidence")

        if is_evidence:
            # --- EVIDENCE: full exhibit treatment (label, cover page, Exhibit prefix) ---
            with get_db() as db:
                exhibit_label, exhibit_order = _next_exhibit_label(db, case_id, numbering)
                exhibit_name = _ai_generate_exhibit_name(text, original_filename)

                # Stamp PDF with cover page
                if ext == ".pdf":
                    stamped_bytes = _stamp_exhibit_on_pdf(file_bytes, exhibit_label, exhibit_name)
                    file_path_on_disk.write_bytes(stamped_bytes)

                new_filename = f"Exhibit {exhibit_label} — {exhibit_name}{ext}"
                db.execute(
                    "UPDATE documents SET exhibit_label = ?, exhibit_name = ?, exhibit_order = ?, filename = ? WHERE id = ?",
                    (exhibit_label, exhibit_name, exhibit_order, new_filename, doc_id)
                )
                logger.info(f"Exhibit processed: {doc_id} -> {new_filename}")
        else:
            # --- NON-EVIDENCE: just AI-rename, no exhibit label or stamp ---
            doc_name = _ai_generate_document_name(text, original_filename)
            new_filename = f"{doc_name}{ext}"
            with get_db() as db:
                db.execute(
                    "UPDATE documents SET filename = ? WHERE id = ?",
                    (new_filename, doc_id)
                )
                logger.info(f"Document renamed: {doc_id} -> {new_filename}")

    except Exception as e:
        logger.error(f"Document processing failed for {doc_id}: {e}\n{traceback.format_exc()}")
        # Fallback for evidence: still try to assign a label
        if category == "evidence":
            try:
                with get_db() as db:
                    exhibit_label, exhibit_order = _next_exhibit_label(db, case_id, numbering)
                    fallback_name = Path(original_filename).stem.replace("_", " ").replace("-", " ")
                    db.execute(
                        "UPDATE documents SET exhibit_label = ?, exhibit_name = ?, exhibit_order = ? WHERE id = ?",
                        (exhibit_label, fallback_name, exhibit_order, doc_id)
                    )
            except Exception:
                pass


# Pydantic models for exhibit endpoints
class ExhibitUpdate(BaseModel):
    exhibit_name: Optional[str] = None
    exhibit_label: Optional[str] = None

class DocumentRename(BaseModel):
    filename: str

class DocumentCategoryUpdate(BaseModel):
    category: str

class ExhibitReorderItem(BaseModel):
    doc_id: str
    order: int

class ExhibitReorderRequest(BaseModel):
    items: List[ExhibitReorderItem]

class DocumentMergeRequest(BaseModel):
    doc_ids: List[str] = []  # empty = merge all
    merge_all: bool = False
    filename: str = "Merged Document"
    category: str = "evidence"
    bates_prefix: str = ""  # e.g. "BATES" — empty = no Bates numbering
    bates_start: int = 1

router = APIRouter(prefix="/api/cases", tags=["cases"])


def _calculate_urgency_score(case: dict) -> float:
    """Calculate urgency score based on deadline proximity vs completion percentage."""
    score = 50.0  # base score

    if case.get("filing_deadline"):
        try:
            deadline = datetime.fromisoformat(case["filing_deadline"])
            now = datetime.now(timezone.utc)
            days_remaining = (deadline - now).days
            if days_remaining < 0:
                score += 50  # Overdue
            elif days_remaining < 3:
                score += 40
            elif days_remaining < 7:
                score += 30
            elif days_remaining < 14:
                score += 20
            elif days_remaining < 30:
                score += 10
        except (ValueError, TypeError):
            pass

    completion = case.get("completion_percentage", 0) or 0
    score -= completion * 0.3  # Higher completion reduces urgency

    priority_weights = {"critical": 20, "high": 15, "medium": 5, "low": 0}
    score += priority_weights.get(case.get("priority", "medium"), 5)

    return max(0, min(100, score))


@router.get("")
async def list_cases(
    status: str = None,
    case_type: str = None,
    sort_by: str = "urgency_score",
    current_user: dict = Depends(get_current_user)
):
    """List cases filtered by tenant."""
    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        query = "SELECT * FROM cases WHERE tenant_id = ?"
        params: list = [tenant_id]
        if status:
            query += " AND status = ?"
            params.append(status)
        if case_type:
            query += " AND case_type = ?"
            params.append(case_type)

        if sort_by == "urgency_score":
            query += " ORDER BY urgency_score DESC"
        elif sort_by == "filing_deadline":
            query += " ORDER BY filing_deadline ASC"
        elif sort_by == "created_at":
            query += " ORDER BY created_at DESC"
        else:
            query += " ORDER BY urgency_score DESC"

        cases = db.execute(query, params).fetchall()
        result = []
        for c in cases:
            case_dict = dict(c)
            # Recalculate urgency score
            new_score = _calculate_urgency_score(case_dict)
            if abs(new_score - (case_dict.get("urgency_score") or 0)) > 1:
                db.execute(
                    "UPDATE cases SET urgency_score = ? WHERE id = ?",
                    (new_score, case_dict["id"])
                )
                case_dict["urgency_score"] = new_score

            # Get task stats
            tasks = db.execute(
                "SELECT status, COUNT(*) as cnt FROM tasks WHERE case_id = ? GROUP BY status",
                (case_dict["id"],)
            ).fetchall()
            task_stats = {t["status"]: t["cnt"] for t in tasks}
            case_dict["task_stats"] = task_stats
            total = sum(task_stats.values())
            completed = task_stats.get("completed", 0)
            if total > 0:
                case_dict["completion_percentage"] = round((completed / total) * 100, 1)

            result.append(case_dict)
        return result


@router.post("")
async def create_case(req: CaseCreate, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Create a new case with AI-generated tasks."""
    tenant_id = current_user["tenant_id"]
    case_id = generate_id()
    now = datetime.now(timezone.utc).isoformat()

    case_number = (req.case_number or "").strip() or None
    exhibit_numbering = req.exhibit_numbering if req.exhibit_numbering in ("letters", "numbers") else "letters"

    # Derive case_type from forum/jurisdiction for backward compat with workflow templates
    derived_case_type = req.case_type or "other"
    if req.forum:
        forum_lower = req.forum.lower()
        if any(x in forum_lower for x in ["uscis", "eoir", "bia", "immigration", "asylum", "dhs", "ice", "cbp", "nvc", "dos"]):
            derived_case_type = "immigration"
        elif any(x in forum_lower for x in ["arbitration"]):
            derived_case_type = "arbitration"
        elif any(x in forum_lower for x in ["mediation"]):
            derived_case_type = "mediation"
        elif any(x in forum_lower for x in ["criminal", "crown court", "magistrate"]):
            derived_case_type = "criminal"
        elif any(x in forum_lower for x in ["family"]):
            derived_case_type = "family"

    # Prefer court from party_roles / request, fallback to forum
    court_value = req.court or req.forum or None

    case_data = {
        "id": case_id,
        "tenant_id": tenant_id,
        "title": req.title,
        "case_number": case_number,
        "case_type": derived_case_type,
        "jurisdiction": req.jurisdiction,
        "forum": req.forum,
        "matter_type": req.matter_type,
        "party_roles": req.party_roles,
        "description": req.description,
        "client_name": req.client_name,
        "opposing_party": req.opposing_party,
        "court": court_value,
        "judge": req.judge,
        "filing_deadline": req.filing_deadline,
        "trial_date": req.trial_date,
        "uscis_receipt_number": req.uscis_receipt_number,
        "priority": req.priority,
        "exhibit_numbering": exhibit_numbering,
        "created_by": current_user["sub"],
        "assigned_attorney_id": current_user["sub"],
        "created_at": now,
        "updated_at": now,
    }

    urgency = _calculate_urgency_score(case_data)
    case_data["urgency_score"] = urgency

    # --- Try AI task generation first ---
    ai_tasks = _ai_generate_case_tasks(
        jurisdiction=req.jurisdiction or "",
        forum=req.forum or "",
        matter_type=req.matter_type or req.case_type or "",
        case_title=req.title,
    )
    ai_tasks_generated = len(ai_tasks) > 0

    with get_db() as db:
        db.execute(
            """INSERT INTO cases (id, tenant_id, title, case_number, case_type,
               jurisdiction, forum, matter_type, party_roles,
               description, client_name, opposing_party, court, judge,
               filing_deadline, trial_date, uscis_receipt_number,
               priority, urgency_score, exhibit_numbering,
               created_by, assigned_attorney_id, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (case_id, tenant_id, req.title, case_number, derived_case_type,
             req.jurisdiction, req.forum, req.matter_type, req.party_roles,
             req.description, req.client_name, req.opposing_party, court_value, req.judge,
             req.filing_deadline, req.trial_date, req.uscis_receipt_number,
             req.priority, urgency, exhibit_numbering,
             current_user["sub"], current_user["sub"], now, now)
        )

        if ai_tasks_generated:
            # Insert AI-generated tasks
            for task_title in ai_tasks:
                task_id = generate_id()
                db.execute(
                    """INSERT INTO tasks (id, case_id, tenant_id, title, status, priority)
                       VALUES (?, ?, ?, ?, 'pending', 'medium')""",
                    (task_id, case_id, tenant_id, task_title)
                )
        else:
            # Fallback: use workflow template if available
            template = db.execute(
                "SELECT * FROM workflow_templates WHERE case_type = ?",
                (derived_case_type,)
            ).fetchone()
            if template:
                tasks = json.loads(template["tasks_json"])
                for task_title in tasks:
                    task_id = generate_id()
                    db.execute(
                        """INSERT INTO tasks (id, case_id, tenant_id, title, status, priority)
                           VALUES (?, ?, ?, ?, 'pending', 'medium')""",
                        (task_id, case_id, tenant_id, task_title)
                    )

    case_data["task_stats"] = {}
    case_data["ai_tasks_generated"] = ai_tasks_generated
    case_data["task_count"] = len(ai_tasks)
    return case_data


@router.get("/{case_id}")
async def get_case(case_id: str, current_user: dict = Depends(get_current_user)):
    """Get case details with tasks, documents, and timeline."""
    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        case = db.execute(
            "SELECT * FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            # Check if user has expert access
            access = db.execute(
                """SELECT ce.* FROM case_experts ce
                   JOIN cases c ON ce.case_id = c.id
                   WHERE ce.case_id = ? AND ce.expert_id = ? AND ce.status = 'active'""",
                (case_id, current_user["sub"])
            ).fetchone()
            if not access:
                raise HTTPException(status_code=404, detail="Case not found")
            case = db.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone()

        case_dict = dict(case)

        # Get tasks
        tasks = db.execute(
            "SELECT * FROM tasks WHERE case_id = ? ORDER BY created_at ASC",
            (case_id,)
        ).fetchall()
        case_dict["tasks"] = [dict(t) for t in tasks]

        # Get documents (ordered by exhibit_order, then created_at)
        docs = db.execute(
            "SELECT * FROM documents WHERE case_id = ? ORDER BY exhibit_order ASC, created_at ASC",
            (case_id,)
        ).fetchall()
        case_dict["documents"] = [dict(d) for d in docs]

        # Get timeline events
        timeline = db.execute(
            "SELECT * FROM case_timeline WHERE case_id = ? ORDER BY event_date ASC",
            (case_id,)
        ).fetchall()
        case_dict["timeline"] = [dict(t) for t in timeline]

        # Get contradictions
        contradictions = db.execute(
            "SELECT * FROM contradictions WHERE case_id = ? ORDER BY severity DESC",
            (case_id,)
        ).fetchall()
        case_dict["contradictions"] = [dict(c) for c in contradictions]

        # Get assigned experts
        experts = db.execute(
            """SELECT ce.*, u.full_name, u.email, u.specializations
               FROM case_experts ce JOIN users u ON ce.expert_id = u.id
               WHERE ce.case_id = ? AND ce.status = 'active'""",
            (case_id,)
        ).fetchall()
        case_dict["experts"] = [dict(e) for e in experts]

        return case_dict


@router.patch("/{case_id}")
async def update_case(
    case_id: str,
    req: CaseUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update case details."""
    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        case = db.execute(
            "SELECT * FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        updates = {}
        nullable_fields = {
            "case_number",
            "description",
            "client_name",
            "opposing_party",
            "court",
            "judge",
            "filing_deadline",
            "trial_date",
            "uscis_receipt_number",
            "assigned_attorney_id",
            "exhibit_numbering",
        }
        for field, value in req.model_dump(exclude_unset=True).items():
            # Allow explicitly clearing nullable fields by sending null
            if value is not None or field in nullable_fields:
                updates[field] = value

        if not updates:
            return dict(case)

        updates["updated_at"] = datetime.now(timezone.utc).isoformat()

        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        values = list(updates.values()) + [case_id, tenant_id]
        db.execute(
            f"UPDATE cases SET {set_clause} WHERE id = ? AND tenant_id = ?",
            values
        )

        # Recalculate urgency
        updated = db.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone()
        updated_dict = dict(updated)
        new_score = _calculate_urgency_score(updated_dict)
        db.execute("UPDATE cases SET urgency_score = ? WHERE id = ?", (new_score, case_id))
        updated_dict["urgency_score"] = new_score

        return updated_dict


@router.delete("/{case_id}")
async def delete_case(case_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a case and all associated data."""
    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        case = db.execute(
            "SELECT * FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        # Delete uploaded files from disk (best effort)
        docs = db.execute(
            "SELECT file_path FROM documents WHERE case_id = ?",
            (case_id,)
        ).fetchall()
        for d in docs:
            try:
                fp = d["file_path"]
                if fp and os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                # Don't block deletion if file cleanup fails
                pass

        # Delete drafts + dependent records
        draft_rows = db.execute(
            "SELECT id FROM legal_drafts WHERE case_id = ?",
            (case_id,)
        ).fetchall()
        draft_ids = [r["id"] for r in draft_rows]
        for did in draft_ids:
            db.execute("DELETE FROM draft_versions WHERE draft_id = ?", (did,))
            db.execute("DELETE FROM research_citations WHERE draft_id = ?", (did,))
            db.execute("DELETE FROM draft_comments WHERE draft_id = ?", (did,))
        db.execute("DELETE FROM legal_drafts WHERE case_id = ?", (case_id,))

        # Delete all associated data
        db.execute("DELETE FROM tasks WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM time_entries WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM documents WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM case_timeline WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM contradictions WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM case_experts WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM waitlist WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM discovery_items WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM witnesses WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM chat_messages WHERE case_id = ?", (case_id,))
        db.execute("DELETE FROM cases WHERE id = ? AND tenant_id = ?", (case_id, tenant_id))

    return {"ok": True, "message": "Case deleted"}


# Task sub-routes
@router.get("/{case_id}/tasks")
async def get_tasks(case_id: str, current_user: dict = Depends(get_current_user)):
    """Get all tasks for a case."""
    with get_db() as db:
        tasks = db.execute(
            "SELECT * FROM tasks WHERE case_id = ? ORDER BY created_at ASC",
            (case_id,)
        ).fetchall()
        return [dict(t) for t in tasks]


@router.post("/{case_id}/tasks")
async def create_task(
    case_id: str,
    req: TaskCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a task for a case."""
    tenant_id = current_user["tenant_id"]
    task_id = generate_id()

    with get_db() as db:
        # Verify case exists
        case = db.execute(
            "SELECT id FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        db.execute(
            """INSERT INTO tasks (id, case_id, tenant_id, title, description, assigned_to,
               due_date, priority, parent_task_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (task_id, case_id, tenant_id, req.title, req.description, req.assigned_to,
             req.due_date, req.priority, req.parent_task_id)
        )
        return {
            "id": task_id, "case_id": case_id, "title": req.title,
            "status": "pending", "priority": req.priority
        }


@router.patch("/tasks/{task_id}")
async def update_task(
    task_id: str,
    req: TaskUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update task status/details."""
    with get_db() as db:
        task = db.execute(
            "SELECT * FROM tasks WHERE id = ? AND tenant_id = ?",
            (task_id, current_user["tenant_id"])
        ).fetchone()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")

        updates = {}
        for field, value in req.model_dump(exclude_unset=True).items():
            if value is not None:
                updates[field] = value

        if "status" in updates and updates["status"] == "completed":
            updates["completed_at"] = datetime.now(timezone.utc).isoformat()

        if not updates:
            return dict(task)

        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        values = list(updates.values()) + [task_id]
        db.execute(f"UPDATE tasks SET {set_clause} WHERE id = ?", values)

        updated = db.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
        return dict(updated)


# Document sub-routes
@router.get("/{case_id}/documents")
async def get_documents(case_id: str, current_user: dict = Depends(get_current_user)):
    """List documents for a case."""
    with get_db() as db:
        docs = db.execute(
            "SELECT * FROM documents WHERE case_id = ? ORDER BY created_at DESC",
            (case_id,)
        ).fetchall()
        return [dict(d) for d in docs]


@router.post("/{case_id}/documents")
async def create_document(
    case_id: str,
    req: DocumentCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a document record (file upload handled separately via pre-signed URL)."""
    tenant_id = current_user["tenant_id"]
    doc_id = generate_id()
    file_path = f"tenants/{tenant_id}/cases/{case_id}/{doc_id}/{req.filename}"

    with get_db() as db:
        db.execute(
            """INSERT INTO documents (id, case_id, tenant_id, filename, file_path, category,
               uploaded_by, content_text)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (doc_id, case_id, tenant_id, req.filename, file_path, req.category,
             current_user["sub"], req.content_text)
        )
        return {
            "id": doc_id,
            "case_id": case_id,
            "filename": req.filename,
            "file_path": file_path,
            "category": req.category,
        }


# File upload endpoint
@router.post("/{case_id}/documents/upload")
async def upload_document_file(
    case_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    category: str = Form(default="general"),
    notes: str = Form(default=""),
    current_user: dict = Depends(get_current_user)
):
    """Upload an actual file to a case. Auto-assigns exhibit label and AI-generated name."""
    tenant_id = current_user["tenant_id"]

    # Validate file extension
    filename = file.filename or "unnamed_file"
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' not allowed. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}")

    # Read file content
    file_bytes = await file.read()
    file_size = len(file_bytes)
    if file_size > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB")
    if file_size == 0:
        raise HTTPException(status_code=400, detail="File is empty")

    doc_id = generate_id()
    mime_type = file.content_type or "application/octet-stream"

    # Create directory and save file
    upload_dir = Path(UPLOAD_BASE_DIR) / tenant_id / case_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = f"{doc_id}_{filename}"
    file_path = upload_dir / safe_filename
    file_path.write_bytes(file_bytes)

    # Store relative path in DB
    relative_path = f"{tenant_id}/{case_id}/{safe_filename}"

    with get_db() as db:
        # Verify case exists
        case = db.execute(
            "SELECT id FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            # Clean up file
            file_path.unlink(missing_ok=True)
            raise HTTPException(status_code=404, detail="Case not found")

        db.execute(
            """INSERT INTO documents (id, case_id, tenant_id, filename, file_path, file_size, mime_type,
               category, uploaded_by, content_text)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (doc_id, case_id, tenant_id, filename, relative_path, file_size, mime_type,
             category, current_user["sub"], notes)
        )

    # Process document in background (AI naming; exhibit labeling only for evidence)
    background_tasks.add_task(
        _process_exhibit_async, doc_id, case_id, tenant_id,
        file_bytes, filename, ext, file_path, relative_path,
        category
    )

    return {
        "id": doc_id,
        "case_id": case_id,
        "filename": filename,
        "file_size": file_size,
        "mime_type": mime_type,
        "category": category,
        "has_file": True,
        "exhibit_processing": category == "evidence",
    }


# File download endpoint
@router.get("/documents/{doc_id}/download")
async def download_document_file(
    doc_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download a document file."""
    with get_db() as db:
        doc = db.execute(
            "SELECT * FROM documents WHERE id = ? AND tenant_id = ?",
            (doc_id, current_user["tenant_id"])
        ).fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        file_path = Path(UPLOAD_BASE_DIR) / doc["file_path"]
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found on server")

        return FileResponse(
            path=str(file_path),
            filename=doc["filename"],
            media_type=doc["mime_type"] or "application/octet-stream",
        )


@router.get("/{case_id}/documents/download-all")
async def download_all_documents_as_pdf(
    case_id: str,
    bates_prefix: str = "",
    bates_start: int = 1,
    current_user: dict = Depends(get_current_user)
):
    """Merge ALL case documents (every file type) into one PDF for download.
    Optional Bates numbering via query params: ?bates_prefix=BATES&bates_start=1"""
    import fitz
    from fastapi.responses import Response

    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        case = db.execute(
            "SELECT * FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        docs = db.execute(
            "SELECT * FROM documents WHERE case_id = ? AND tenant_id = ? ORDER BY exhibit_order ASC, created_at ASC",
            (case_id, tenant_id)
        ).fetchall()
        if not docs:
            raise HTTPException(status_code=404, detail="No documents in this case")
        docs = [dict(d) for d in docs]

    # Sort: evidence/exhibits by exhibit_label (A, B, C...), then court_filing, then everything else by date
    # This ensures Exhibit A always comes first, then B, C, etc.
    def _doc_sort_key(d):
        label = d.get('exhibit_label') or ''
        cat = d.get('category', '')
        if cat == 'evidence' and label:
            # Evidence with exhibit label: sort first, by label alphabetically
            return (0, label, d.get('created_at', ''))
        elif cat == 'court_filing':
            return (1, '', d.get('created_at', ''))
        elif cat == 'petition':
            return (2, '', d.get('created_at', ''))
        elif cat == 'evidence':
            # Evidence without label yet: after labeled evidence
            return (0, 'ZZZZ', d.get('created_at', ''))
        else:
            return (3, '', d.get('created_at', ''))
    docs.sort(key=_doc_sort_key)

    merged_pdf = fitz.open()
    for doc in docs:
        if not doc.get("file_path"):
            continue
        fp = Path(UPLOAD_BASE_DIR) / doc["file_path"]
        if not fp.exists():
            continue
        _add_file_to_pdf(merged_pdf, fp, doc.get("filename", ""))

    if merged_pdf.page_count == 0:
        merged_pdf.close()
        raise HTTPException(status_code=400, detail="No mergeable content found")

    # Apply Bates numbering if requested
    if bates_prefix:
        _apply_bates_numbering(merged_pdf, bates_prefix, bates_start)

    merged_bytes = merged_pdf.tobytes()
    merged_pdf.close()

    case_title = dict(case).get("title", "Case").replace(" ", "_")
    filename = f"{case_title}_Complete_Filing.pdf"

    return Response(
        content=merged_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{case_id}/documents/download-zip")
async def download_all_documents_as_zip(
    case_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download ALL case documents as a ZIP archive, preserving original file formats."""
    import zipfile
    from fastapi.responses import Response

    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        case = db.execute(
            "SELECT * FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        docs = db.execute(
            "SELECT * FROM documents WHERE case_id = ? AND tenant_id = ? ORDER BY exhibit_order ASC, created_at ASC",
            (case_id, tenant_id)
        ).fetchall()
        if not docs:
            raise HTTPException(status_code=404, detail="No documents in this case")
        docs = [dict(d) for d in docs]

    # Sort by exhibit label for consistent ordering
    def _zip_sort_key(d):
        label = d.get('exhibit_label') or ''
        cat = d.get('category', '')
        if cat == 'evidence' and label:
            return (0, label, d.get('created_at', ''))
        elif cat == 'court_filing':
            return (1, '', d.get('created_at', ''))
        elif cat == 'petition':
            return (2, '', d.get('created_at', ''))
        elif cat == 'evidence':
            return (0, 'ZZZZ', d.get('created_at', ''))
        else:
            return (3, '', d.get('created_at', ''))
    docs.sort(key=_zip_sort_key)

    buf = io.BytesIO()
    used_names: dict = {}
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for doc in docs:
            if not doc.get("file_path"):
                continue
            fp = Path(UPLOAD_BASE_DIR) / doc["file_path"]
            if not fp.exists():
                continue
            # Use display filename, deduplicate if needed
            name = doc.get("filename") or fp.name
            if name in used_names:
                used_names[name] += 1
                stem = Path(name).stem
                ext = Path(name).suffix
                name = f"{stem} ({used_names[name]}){ext}"
            else:
                used_names[name] = 1
            zf.write(str(fp), name)

    zip_bytes = buf.getvalue()
    case_title = dict(case).get("title", "Case").replace(" ", "_")
    filename = f"{case_title}_Documents.zip"

    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# Delete document endpoint
@router.delete("/documents/{doc_id}")
async def delete_document(
    doc_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a document and its file."""
    with get_db() as db:
        doc = db.execute(
            "SELECT * FROM documents WHERE id = ? AND tenant_id = ?",
            (doc_id, current_user["tenant_id"])
        ).fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        # Delete file from disk
        if doc["file_path"]:
            file_path = Path(UPLOAD_BASE_DIR) / doc["file_path"]
            file_path.unlink(missing_ok=True)

        db.execute("DELETE FROM documents WHERE id = ?", (doc_id,))

    return {"status": "deleted", "document_id": doc_id}


@router.patch("/documents/{doc_id}/rename")
async def rename_document(
    doc_id: str,
    req: DocumentRename,
    current_user: dict = Depends(get_current_user)
):
    """Rename a document's display filename."""
    new_name = (req.filename or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Filename cannot be empty")

    with get_db() as db:
        doc = db.execute(
            "SELECT * FROM documents WHERE id = ? AND tenant_id = ?",
            (doc_id, current_user["tenant_id"])
        ).fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        db.execute("UPDATE documents SET filename = ? WHERE id = ?", (new_name, doc_id))
        updated = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
        return dict(updated)


@router.patch("/documents/{doc_id}/category")
async def update_document_category(
    doc_id: str,
    req: DocumentCategoryUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a document's category (e.g. move to 'ready' folder)."""
    valid_categories = {'general', 'evidence', 'petition', 'correspondence', 'court_filing', 'ready'}
    cat = (req.category or "").strip().lower()
    if cat not in valid_categories:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {', '.join(sorted(valid_categories))}")

    with get_db() as db:
        doc = db.execute(
            "SELECT * FROM documents WHERE id = ? AND tenant_id = ?",
            (doc_id, current_user["tenant_id"])
        ).fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        db.execute("UPDATE documents SET category = ? WHERE id = ?", (cat, doc_id))
        updated = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
        return dict(updated)


# --------------- Document merge endpoint ---------------

@router.post("/{case_id}/documents/merge")
async def merge_documents(
    case_id: str,
    req: DocumentMergeRequest,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Merge documents into a single PDF. Supports ALL file types.
    If merge_all=true, merges every document in the case. Otherwise merges doc_ids (1+)."""
    import fitz

    tenant_id = current_user["tenant_id"]

    with get_db() as db:
        # Verify case
        case = db.execute(
            "SELECT id FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        if req.merge_all:
            # Merge ALL documents in the case
            rows = db.execute(
                "SELECT * FROM documents WHERE case_id = ? AND tenant_id = ? ORDER BY exhibit_order ASC, created_at ASC",
                (case_id, tenant_id)
            ).fetchall()
            if not rows:
                raise HTTPException(status_code=400, detail="No documents in this case to merge")
            docs = [dict(r) for r in rows]
            # Sort like download-all: exhibits by label (A, B, C...), then other categories
            def _merge_sort_key(d):
                label = d.get('exhibit_label') or ''
                cat = d.get('category', '')
                if cat == 'evidence' and label:
                    return (0, label, d.get('created_at', ''))
                elif cat == 'court_filing':
                    return (1, '', d.get('created_at', ''))
                elif cat == 'petition':
                    return (2, '', d.get('created_at', ''))
                elif cat == 'evidence':
                    return (0, 'ZZZZ', d.get('created_at', ''))
                else:
                    return (3, '', d.get('created_at', ''))
            docs.sort(key=_merge_sort_key)
        else:
            if len(req.doc_ids) < 1:
                raise HTTPException(status_code=400, detail="Select at least 1 document to merge")
            # Fetch docs in the order requested
            docs = []
            for did in req.doc_ids:
                doc = db.execute(
                    "SELECT * FROM documents WHERE id = ? AND case_id = ? AND tenant_id = ?",
                    (did, case_id, tenant_id)
                ).fetchone()
                if not doc:
                    raise HTTPException(status_code=404, detail=f"Document {did} not found")
                docs.append(dict(doc))

    # Merge all file types using shared helper
    merged_pdf = fitz.open()
    for doc in docs:
        if not doc.get("file_path"):
            continue
        fp = Path(UPLOAD_BASE_DIR) / doc["file_path"]
        if not fp.exists():
            continue
        _add_file_to_pdf(merged_pdf, fp, doc.get("filename", ""))

    if merged_pdf.page_count == 0:
        merged_pdf.close()
        raise HTTPException(status_code=400, detail="No mergeable content found in selected documents")

    # Apply Bates numbering if requested
    if req.bates_prefix:
        _apply_bates_numbering(merged_pdf, req.bates_prefix, req.bates_start)

    # Save merged file
    merged_bytes = merged_pdf.tobytes()
    merged_pdf.close()

    doc_id = generate_id()
    upload_dir = Path(UPLOAD_BASE_DIR) / tenant_id / case_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = f"{doc_id}_merged.pdf"
    file_path = upload_dir / safe_filename
    file_path.write_bytes(merged_bytes)
    relative_path = f"{tenant_id}/{case_id}/{safe_filename}"

    merged_count = len(docs)
    display_name = (req.filename or "Merged Document").strip()
    if not display_name.lower().endswith(".pdf"):
        display_name += ".pdf"

    with get_db() as db:
        db.execute(
            """INSERT INTO documents (id, case_id, tenant_id, filename, file_path, file_size, mime_type,
               category, uploaded_by, content_text, is_merged)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
            (doc_id, case_id, tenant_id, display_name, relative_path, len(merged_bytes),
             "application/pdf", req.category, current_user["sub"],
             f"Merged from {merged_count} documents" + (f" | Bates: {req.bates_prefix}-{req.bates_start:06d}" if req.bates_prefix else ""))
        )

    # Process in background (exhibit treatment if evidence, AI rename otherwise)
    background_tasks.add_task(
        _process_exhibit_async, doc_id, case_id, tenant_id,
        merged_bytes, display_name, ".pdf", file_path, relative_path,
        req.category
    )

    return {
        "id": doc_id,
        "case_id": case_id,
        "filename": display_name,
        "file_size": len(merged_bytes),
        "category": req.category,
        "merged_from": [d["id"] for d in docs],
        "merged_count": merged_count,
        "bates_applied": bool(req.bates_prefix),
        "exhibit_processing": req.category == "evidence",
    }


# --------------- Exhibit management endpoints ---------------

@router.patch("/documents/{doc_id}/exhibit")
async def update_exhibit(
    doc_id: str,
    req: ExhibitUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Manually update exhibit name or label for a document."""
    with get_db() as db:
        doc = db.execute(
            "SELECT * FROM documents WHERE id = ? AND tenant_id = ?",
            (doc_id, current_user["tenant_id"])
        ).fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        updates = {}
        if req.exhibit_name is not None:
            updates["exhibit_name"] = req.exhibit_name
        if req.exhibit_label is not None:
            updates["exhibit_label"] = req.exhibit_label

        if not updates:
            return dict(doc)

        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        values = list(updates.values()) + [doc_id]
        db.execute(f"UPDATE documents SET {set_clause} WHERE id = ?", values)

        # If label or name changed, update filename and re-stamp PDF
        if req.exhibit_label is not None or req.exhibit_name is not None:
            updated_doc = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
            updated_doc = dict(updated_doc)
            label = updated_doc.get("exhibit_label", "")
            name = updated_doc.get("exhibit_name", "")
            # Determine extension from file_path
            ext = Path(updated_doc.get("file_path", "")).suffix.lower() if updated_doc.get("file_path") else ".pdf"
            new_filename = f"Exhibit {label} — {name}{ext}"
            db.execute("UPDATE documents SET filename = ? WHERE id = ?", (new_filename, doc_id))

            if updated_doc.get("file_path") and updated_doc.get("mime_type", "").startswith("application/pdf"):
                try:
                    fp = Path(UPLOAD_BASE_DIR) / updated_doc["file_path"]
                    if fp.exists():
                        original_bytes = fp.read_bytes()
                        stamped = _stamp_exhibit_on_pdf(original_bytes, label, name)
                        fp.write_bytes(stamped)
                except Exception as e:
                    logger.warning(f"Re-stamp failed: {e}")

        updated = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
        return dict(updated)


@router.post("/{case_id}/documents/reorder")
async def reorder_exhibits(
    case_id: str,
    req: ExhibitReorderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Reorder exhibits for a case. Reassigns labels A, B, C... based on new order."""
    tenant_id = current_user["tenant_id"]
    with get_db() as db:
        # Verify case
        case = db.execute(
            "SELECT id FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, tenant_id)
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")

        # Sort items by requested order
        sorted_items = sorted(req.items, key=lambda x: x.order)

        for i, item in enumerate(sorted_items):
            new_label = _exhibit_letter(i)
            doc = db.execute("SELECT * FROM documents WHERE id = ?", (item.doc_id,)).fetchone()
            exhibit_name = doc["exhibit_name"] if doc else ""
            ext = Path(doc["file_path"]).suffix.lower() if doc and doc.get("file_path") else ".pdf"
            new_filename = f"Exhibit {new_label} — {exhibit_name}{ext}"

            db.execute(
                "UPDATE documents SET exhibit_order = ?, exhibit_label = ?, filename = ? WHERE id = ? AND case_id = ?",
                (i, new_label, new_filename, item.doc_id, case_id)
            )

            # Re-stamp PDFs with new label
            if doc and doc["file_path"] and (doc.get("mime_type") or "").startswith("application/pdf"):
                try:
                    fp = Path(UPLOAD_BASE_DIR) / doc["file_path"]
                    if fp.exists():
                        original_bytes = fp.read_bytes()
                        stamped = _stamp_exhibit_on_pdf(
                            original_bytes, new_label, exhibit_name
                        )
                        fp.write_bytes(stamped)
                except Exception as e:
                    logger.warning(f"Re-stamp on reorder failed for {item.doc_id}: {e}")

        # Return updated documents
        docs = db.execute(
            "SELECT * FROM documents WHERE case_id = ? ORDER BY exhibit_order ASC, created_at ASC",
            (case_id,)
        ).fetchall()
        return [dict(d) for d in docs]


@router.post("/documents/{doc_id}/retrigger-exhibit-ai")
async def retrigger_exhibit_ai(
    doc_id: str,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Re-trigger AI naming for a document."""
    with get_db() as db:
        doc = db.execute(
            "SELECT * FROM documents WHERE id = ? AND tenant_id = ?",
            (doc_id, current_user["tenant_id"])
        ).fetchone()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")

        doc = dict(doc)
        if not doc.get("file_path"):
            raise HTTPException(status_code=400, detail="No file associated with this document")

        fp = Path(UPLOAD_BASE_DIR) / doc["file_path"]
        if not fp.exists():
            raise HTTPException(status_code=404, detail="File not found on server")

        file_bytes = fp.read_bytes()
        ext = Path(doc["filename"]).suffix.lower()

        background_tasks.add_task(
            _retrigger_ai_task, doc_id, doc["case_id"], doc["tenant_id"],
            file_bytes, doc["filename"], ext, fp
        )

    return {"status": "processing", "doc_id": doc_id}


def _retrigger_ai_task(doc_id: str, case_id: str, tenant_id: str,
                       file_bytes: bytes, filename: str, ext: str,
                       file_path_on_disk: Path):
    """Background: re-run AI naming and re-stamp PDF."""
    try:
        if ext == ".pdf":
            text = _extract_text_from_pdf(file_bytes)
        elif ext in (".docx", ".doc"):
            text = _extract_text_from_docx(file_bytes)
        elif ext == ".txt":
            text = file_bytes.decode("utf-8", errors="ignore")[:3000]
        else:
            text = ""

        exhibit_name = _ai_generate_exhibit_name(text, filename)

        with get_db() as db:
            doc = db.execute("SELECT exhibit_label, file_path FROM documents WHERE id = ?", (doc_id,)).fetchone()
            exhibit_label = doc["exhibit_label"] if doc else "A"
            file_ext = Path(doc["file_path"]).suffix.lower() if doc and doc.get("file_path") else ext
            new_filename = f"Exhibit {exhibit_label} — {exhibit_name}{file_ext}"

            db.execute(
                "UPDATE documents SET exhibit_name = ?, filename = ? WHERE id = ?",
                (exhibit_name, new_filename, doc_id)
            )

            if ext == ".pdf":
                stamped = _stamp_exhibit_on_pdf(file_bytes, exhibit_label, exhibit_name)
                file_path_on_disk.write_bytes(stamped)

        logger.info(f"AI re-trigger complete: {doc_id} -> {new_filename}")
    except Exception as e:
        logger.error(f"AI re-trigger failed for {doc_id}: {e}")


# USCIS status sub-route
@router.get("/{case_id}/uscis-status")
async def get_uscis_status(case_id: str, current_user: dict = Depends(get_current_user)):
    """Get USCIS case status for a case."""
    with get_db() as db:
        case = db.execute(
            "SELECT uscis_receipt_number, uscis_status, uscis_last_checked FROM cases WHERE id = ? AND tenant_id = ?",
            (case_id, current_user["tenant_id"])
        ).fetchone()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")
        return dict(case)
